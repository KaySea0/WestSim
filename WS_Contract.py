import tkinter as tk
from tkinter import ttk
import json
import openpyxl
import os
import datetime
import smtplib
from pathlib import Path
from tkinter import filedialog
from tkinter import messagebox
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders
from string import Template
from myimages import *
from settings import *

# shorthand for each month to create PO number
month_init = ["JA","FE","MR","AP","MY","JU","JY","AU","SE","OC","NV","DE"]

# if no preservation method is given / is special mil spec, put "n/a" for method in display
# maybe add way to select multiple contracts that apply to certain PO

class WS_Contract(object):

	def __init__(self):
		self.main_wb = None # main workbook that holds key records
		self.wip_wb = None # contains more detailed information about recently added contracts
		self.dict = None # configuration dictionary that contains reference to workbooks
		
		self.PO_dict = None # list of address / contact info for each company
		self.company_list = None # list of companies that user can select for creating / sending PO
		self.current_company = tk.StringVar() # reference variable for currently selected company
		
		self.wip_dict = None # information pertinent to PO creation for each contract in wip_wb
		self.wip_list = None # list of all contract numbers that are in wip_wb
		self.current_contract = tk.StringVar() # reference variable for currently selected contract
		self.current_contract_num = tk.IntVar() # reference variable for contract's line number in main_wb
		
		self.PO_Vars = None # list of all reference variables for PO creation form
		self.check_var = tk.IntVar() # reference variable for checkbox in PO creation form
		
		self.next_row = 0 # local reference to next line for new contracts to be added
		self.next_ref = 0 # local reference to next main workbook reference number in DLAORDERS
		self.contract_edits = [] # list of all new contracts to be added to main workbook
		self.PO_edits = [] # list of all PO info to be added to main workbook
	
	# # #
	# Method: create_dicts
	# Input: n/a
	# Utility:
	#   Populate PO_dict/company_list with company info from cage code list 
	#   and wip_dict/wip_list with contract info from wip list of recent contracts  
	# # #
	def create_dicts(self):
		
		# open up cage code workbook
		cage_wb = openpyxl.load_workbook(self.dict['cage'])
		cage_ws = cage_wb.active

		# instantiate po_dict/company_list 
		self.PO_dict = {}
		self.company_list = []
		
		# grab info for each company in workbook
		for row in cage_ws.iter_rows(min_row=2, values_only=True):
		
			t_entry = {} # temp entry that will be added to main list
			t_entry['line1'] = row[5] if row[5] != None else "" # vendor name
			t_entry['line2'] = row[6] if row[6] != None else "" # address line 1
			t_entry['line3'] = row[7] if row[7] != None else "" # address line 2
			t_entry['line4'] = row[8] if row[8] != None else "" # phone number
			t_entry['line5'] = row[9] if row[9] != None else "" # attention
			t_entry['email'] = row[4] if row[4] != None else "" # email to send PO to
			
			self.company_list.append(row[2]) # add company name to list for user to select
			self.PO_dict[row[2]] = t_entry # add company info to dictionary, make key the name of the company
		
		# open recent contract workbook
		wip_ws = self.wip_wb.active
		
		# instantiate wip_dict/wip_list 
		self.wip_dict = {}
		self.wip_list = []
		
		# grab info for each contract in workbook
		for row in wip_ws.iter_rows(min_row=2, values_only=True):
		
			t_entry = {} 					# temp entry that will be added to main list
			t_entry['pn'] = row[7] 			# part number
			t_entry['nsn'] = row[4] 		# NSN
			t_entry['description'] = row[5] # general part description
			t_entry['qty'] = row[2] 		# quantity
			
			self.wip_list.append(row[1]) # add contract number to list for user to select
			self.wip_dict[row[1]] = t_entry # add contract info to dictionary, make key the contract number
	
	# # #
	# Method: save_changes
	# Input: n/a
	# Utility:
	#   Make all appropriate changes to main_wb from user's input of new contracts and PO's in one go to avoid need for workbook reparation 
	# # #
	def save_changes(self):

		# if either list has values, then workbook needs to be updated
		if self.contract_edits or self.PO_edits:

			main_ws = self.main_wb['DLAORDERS'] # open main workbook
			
			# if new contracts have been added by the user...
			if self.contract_edits:
				for contract in self.contract_edits: # ...run through each contract and add info to workbook
					
					row_num = contract[0] # row in main workbook that info will be put on
					
					main_ws['A'+row_num] = contract[1] # reference number
					main_ws['G'+row_num] = contract[2] # award date
					main_ws['B'+row_num] = contract[3] # contract number
					main_ws['E'+row_num] = contract[4] # quantity
					main_ws['K'+row_num] = contract[5] # contract total
					main_ws['D'+row_num] = contract[6] # NSN
					main_ws['F'+row_num] = contract[7] # vendor name
					main_ws['H'+row_num] = contract[8] # due date

			# if new POs have been made by the user...
			if self.PO_edits:
				for PO in self.PO_edits: # ...run through each PO and add info to workbook
					
					row_num = PO[0] # row in main workbook that info will be put on
					
					main_ws['I'+row_num] = PO[1] # PO number
					main_ws['L'+row_num] = PO[2] # PO total
			
			# attempt to save main workbook, prompt user to close main workbook if it is still open so save operation can occur
			while True:
				try:
					self.main_wb.save(self.dict['main']) # save main_wb
					break
				except:
					pass
					
				tk.messagebox.showinfo("An error has occurred", "Workbook {} is still open, please close it and press 'Ok' to continue operation.".format(self.dict['main'][self.dict['main'].rfind('/')+1:]))
			
			# clear lists of additions to avoid accidental duplication
			self.contract_edits.clear()
			self.PO_edits.clear()
				
			# save confirmation
			tk.messagebox.showinfo("Changes confirmation", "All changes have been saved to Workbook {}.".format(self.dict['main'][self.dict['main'].rfind('/')+1:]))
	
	# # # 
	# Method: save_PO
	# Input:
	#   window - PO creation window reference that will be closed after method completion
	# Utility:
	#   Open up PO template, populate with information from form, save to selected location, and update main workbook with PO number
	# # #
	def save_PO(self, window):
		
		# prompt user to select location for PO to be saved, save location
		f = filedialog.asksaveasfile(mode='w', initialfile="PurchaseOrder - " + self.PO_Vars[5].get(), filetypes=(("Excel file", "*.xlsx"),))
		
		# open PO template workbook for editing
		po_template = openpyxl.load_workbook('PO_Template.xlsx')
		po_ws = po_template.active
		
		# attempt to open Westsim logo and place it in correct place on template
		try: 
			ws_image = openpyxl.drawing.image.Image('logo.png')
			ws_image.anchor(po_ws.cell('B2'))
			po_ws.add_image(ws_image)
		except:
			pass
		
		po_ws['B10'] = self.PO_Vars[0].get() # vendor name
		po_ws['B11'] = self.PO_Vars[1].get() # address line 1
		po_ws['B12'] = self.PO_Vars[2].get() # address line 2
		po_ws['B13'] = self.PO_Vars[3].get() # phone number
		po_ws['B14'] = self.PO_Vars[4].get() # attention
		po_ws['G10'] = "PO: {}".format(self.PO_Vars[5].get())		# PO number
		po_ws['G12'] = "Quote: {}".format(self.PO_Vars[6].get())	# quote reference number
		po_ws['G13'] = "Delivery: {}".format(self.PO_Vars[7].get()) # delivery time
		po_ws['G14'] = "Terms: {}".format(self.PO_Vars[8].get())	# extra terms
		po_ws['C18'] = "P/N: {}".format(self.PO_Vars[9].get())		# part number
		po_ws['C19'] = "NSN: {}".format(self.PO_Vars[10].get())		# NSN
		po_ws['C20'] = self.PO_Vars[11].get() # general part description
		po_ws['E18'] = int(self.PO_Vars[12].get()) # quantity
		po_ws['G18'] = float(self.PO_Vars[13].get()) # unit price
		
		# if checkbox has been checked, include UPS account number
		if self.check_var.get() == 1:
			po_ws['B46'] = "UPS Ground account no: 2Y642X"
			
		# save completed PO to appropriate location and clean reference file
		po_template.save(f.name+".xlsx")
		f.close()
		os.remove(f.name)
		
		info = []
		info.append(str(self.current_contract_num.get()))
		info.append(self.PO_Vars[5].get())
		info.append(int(self.PO_Vars[12].get()) * float(self.PO_Vars[13].get()))
		
		self.PO_edits.append(info)
		
		window.destroy() # close PO form
		
		# confirmation message that PO was saved successfully
		tk.messagebox.showinfo("PO Created", "{} has been saved.".format(f.name))
	
	# # #
	# Method: process_addition
	# Input:
	#   var_list - list of reference variables for contract addition form
	#   window - contract addition window reference that will be closed after completion
	# Utility:
	#   Populate main_wb and wip_wb with pertinent contract information
	# # #
	def process_addition(self, var_list, window):
		
		# temp list that stores current contract addition's information
		info = []
		info.append(str(self.next_row)) # row in main_wb that contract info will be put into
		
		main_ws = self.main_wb['DLAORDERS']
		info.append(self.next_ref) # reference number [A] 
		info.append(var_list[0].get()) # award date [G]
		info.append(var_list[1].get()) # contract number [B]
		info.append(int(var_list[2].get())) # quantity [E]
		info.append(var_list[3].get()) # contract total [K]
		info.append(var_list[4].get()) # NSN [D]
		info.append(var_list[6].get()) # vendor name [F]
		info.append(var_list[9].get()) # due date [H]
		
		# add contract addition info to overarching list and update relevant reference variables
		self.contract_edits.append(info)
		self.next_row += 1
		self.next_ref += 1
		
		# open wip_wb
		wip_ws = self.wip_wb.active
		next_row = wip_ws.max_row+1
		
		wip_ws['A'+str(next_row)] = var_list[0].get() # award date
		wip_ws['B'+str(next_row)] = var_list[1].get() # contract number
		wip_ws['C'+str(next_row)] = var_list[2].get() # quantity
		wip_ws['D'+str(next_row)] = var_list[3].get() # contract total
		wip_ws['E'+str(next_row)] = var_list[4].get() # NSN
		wip_ws['F'+str(next_row)] = var_list[5].get() # general part description
		wip_ws['G'+str(next_row)] = var_list[6].get() # vendor name
		wip_ws['H'+str(next_row)] = var_list[7].get() # part number
		wip_ws['I'+str(next_row)] = var_list[8].get() # preservation method
		wip_ws['J'+str(next_row)] = var_list[9].get() # due date

		# attempt to save wip workbook, prompt user to close wip workbook if it is still open so save operation can occur
		while True:
			try:
				self.wip_wb.save(self.dict['wip']) # save wip_wb
				break
			except:
				pass
				
			tk.messagebox.showinfo("An error has occurred", "Workbook {} is still open, please close it and press 'Ok' to continue operation.".format(self.dict['wip'][self.dict['wip'].rfind('/')+1:]))
		
		window.destroy() # close contract addition form
		
		# confirmation message that contract info was saved
		tk.messagebox.showinfo("Contract info saved", "Contract {} has been registered. Be sure to save the PDF to the appropriate folder ({}).".format(var_list[1].get(),var_list[6].get()))
	
	# # #
	# Method: create_PO
	# Input: n/a
	# Utility:
	#   Create and open form to create PO
	# # #
	def create_PO(self):
		
		# create main PO form window
		t = tk.Toplevel()
		t.geometry('640x450')
		t.title("PO Creation")
		
		# reset combobox reference vars 
		self.current_company.set("")
		self.current_contract.set("")
		
		# open up main_wb in order to create PO number when necessary
		main_ws = self.main_wb['DLAORDERS']
		
		# instantiate all PO reference variables
		self.PO_Vars = []
		for i in range(0,14):
			temp = tk.StringVar()
			self.PO_Vars.append(temp)
			
		# when company is selected, populate form with pertinent information 
		def company_function(eventObject):
		
			company_info = self.PO_dict[self.current_company.get()] # get reference list for chosen company
			self.PO_Vars[0].set(company_info['line1']) # vendor name
			self.PO_Vars[1].set(company_info['line2']) # address line 1
			self.PO_Vars[2].set(company_info['line3']) # address line 2
			self.PO_Vars[3].set(company_info['line4']) # phone number
			self.PO_Vars[4].set(company_info['line5']) # attention
		
		# when contract is selected, populate form with pertinent information
		def contract_function(eventObject):
			
			contract_info = self.wip_dict[self.current_contract.get()] # get reference list for chosen contract
			
			# # # PO number generation start
			contract_trace = main_ws.max_row # reference var to trace through main_wb
			
			# iterate through main_wb until selected contract is found
			while(main_ws['B'+str(contract_trace)].value != self.current_contract.get()):
				contract_trace -= 1
			
			# get local contract number and set reference to contract's line number in main_wb
			contract_num = main_ws['A'+str(contract_trace)].value
			self.current_contract_num.set(contract_trace)
			
			# continue to iterate through main_wb until current month is found
			while type(main_ws['B'+str(contract_trace)].value) != datetime.datetime:
				contract_trace -= 1
			
			# get date portion of PO number (2 character shorthand for month + last 2 digits of current year)
			current_month = main_ws['B'+str(contract_trace)].value
			po_time = str(month_init[current_month.month-1]) + str(current_month.year%100)
			
			# if local contract number is less than 10, add leading 0
			if contract_num < 10:
				po_time += "0"
			
			# create full PO number
			po_num = po_time + str(contract_num)
			# # # PO number generation end
			
			self.PO_Vars[5].set(po_num)							# PO number
			self.PO_Vars[9].set(contract_info['pn'])			# part number
			self.PO_Vars[10].set(contract_info['nsn'])			# NSN
			self.PO_Vars[11].set(contract_info['description'])	# general part description
			self.PO_Vars[12].set(contract_info['qty'])			# quantity
		
		# dropdown menu for company names
		company_menu = ttk.Combobox(t, textvariable=self.current_company, values=self.company_list)
		company_menu.bind('<<ComboboxSelected>>', company_function)
		company_menu.grid(row=0, column=0, columnspan=2, sticky="new", padx=10, pady=10)
		
		# label for vendor name
		company_label = tk.Label(t, text="Vendor Name:")
		company_label.grid(row=1, column=0, sticky="w", padx=10, pady=10)
		
		# text box for vendor name
		company_entry = tk.Entry(t, width=30, textvariable=self.PO_Vars[0])
		company_entry.grid(row=1, column=1, sticky="w", padx=10, pady=10)
		
		# label for address line 1
		addr1_label = tk.Label(t, text="Address Line 1:")
		addr1_label.grid(row=2, column=0, sticky="w", padx=10, pady=10)
		
		# text box for address line 1
		addr1_entry = tk.Entry(t, width=30, textvariable=self.PO_Vars[1])
		addr1_entry.grid(row=2, column=1, sticky="w", padx=10, pady=10)
		
		# label for address line 2
		addr2_label = tk.Label(t, text="Address Line 2:")
		addr2_label.grid(row=3, column=0, sticky="w", padx=10, pady=10)
		
		# text box for address line 2
		addr2_entry = tk.Entry(t, width=30, textvariable=self.PO_Vars[2])
		addr2_entry.grid(row=3, column=1, sticky="w", padx=10, pady=10)
		
		# label for company phone number
		phone_label = tk.Label(t, text="Phone:")
		phone_label.grid(row=4, column=0, sticky="w", padx=10, pady=10)
		
		# text box for company phone number
		phone_entry = tk.Entry(t, width=30, textvariable=self.PO_Vars[3])
		phone_entry.grid(row=4, column=1, sticky="w", padx=10, pady=10)
		
		# label for company attention
		attention_label = tk.Label(t, text="Attention:")
		attention_label.grid(row=5, column=0, sticky="w", padx=10, pady=10)
		
		# text box for company attention
		attention_entry = tk.Entry(t, width=30, textvariable=self.PO_Vars[4])
		attention_entry.grid(row=5, column=1, sticky="w", padx=10, pady=10)
		
		# checkbox for inclusion of UPS Ground account number
		UPS_check = tk.Checkbutton(t, text="Include UPS Ground account #", variable=self.check_var)
		UPS_check.grid(row=0, column=2, columnspan=2, padx=10, pady=10)
		
		# label for PO number
		poNum_label = tk.Label(t, text="PO #:")
		poNum_label.grid(row=1, column=2, sticky="w", padx=10, pady=10)
		
		# text box for PO number
		poNum_entry = tk.Entry(t, width=10, textvariable=self.PO_Vars[5])
		poNum_entry.grid(row=1, column=3, sticky="w", padx=10, pady=10)
		
		# label for quote reference number
		quote_label = tk.Label(t, text="Quote:")
		quote_label.grid(row=2, column=2, sticky="w", padx=10, pady=10)
		
		# text box for quote reference number
		quote_entry = tk.Entry(t, width=30, textvariable=self.PO_Vars[6])
		quote_entry.grid(row=2, column=3, sticky="w", padx=10, pady=10)
		
		# label for delivery time 
		delivery_label = tk.Label(t, text="Delivery:")
		delivery_label.grid(row=3, column=2, sticky="w", padx=10, pady=10)
		
		# text box for delivery time
		delivery_entry = tk.Entry(t, width=30, textvariable=self.PO_Vars[7])
		delivery_entry.grid(row=3, column=3, sticky="w", padx=10, pady=10)
		
		# label for extra terms
		terms_label = tk.Label(t, text="Terms:")
		terms_label.grid(row=4, column=2, sticky="w", padx=10, pady=10)
		
		# text box for extra terms
		terms_entry = tk.Entry(t, width=30, textvariable=self.PO_Vars[8])
		terms_entry.grid(row=4, column=3, sticky="w", padx=10, pady=10)
		
		# label for unit price
		unit_label = tk.Label(t, text="Unit Price:")
		unit_label.grid(row=5, column=2, sticky="w", padx=10, pady=10)
		
		# text box for unit price
		unit_entry = tk.Entry(t, width=30, textvariable=self.PO_Vars[13])
		unit_entry.grid(row=5, column=3, sticky="w", padx=10, pady=10)
		
		# dropdown menu for contract numbers
		contract_menu = ttk.Combobox(t, textvariable=self.current_contract, values=self.wip_list)
		contract_menu.bind('<<ComboboxSelected>>', contract_function)
		contract_menu.grid(row=6, column=0, columnspan=2, sticky="new", padx=10, pady=10)
		
		# label for part number
		pn_label = tk.Label(t, text="P/N:")
		pn_label.grid(row=7, column=0, sticky="w", padx=10, pady=10)
		
		# text box for part number
		pn_entry = tk.Entry(t, width=30, textvariable=self.PO_Vars[9])
		pn_entry.grid(row=7, column=1, sticky="w", padx=10, pady=10)
		
		# label for NSN
		nsn_label = tk.Label(t, text="NSN:")
		nsn_label.grid(row=8, column=0, sticky="w", padx=10, pady=10)
		
		# text box for NSN
		nsn_entry = tk.Entry(t, width=30, textvariable=self.PO_Vars[10])
		nsn_entry.grid(row=8, column=1, sticky="w", padx=10, pady=10)
		
		# label for general part description
		descr_label = tk.Label(t, text="Part Description:")
		descr_label.grid(row=7, column=2, sticky="w", padx=10, pady=10)
		
		# text box for general part description
		descr_entry = tk.Entry(t, width=30, textvariable=self.PO_Vars[11])
		descr_entry.grid(row=7, column=3, sticky="w", padx=10, pady=10)
		
		# label to quantity
		qty_label = tk.Label(t, text="QTY:")
		qty_label.grid(row=8, column=2, sticky="w", padx=10, pady=10)
		
		# text box for quantity
		qty_entry=tk.Entry(t, width=30, textvariable=self.PO_Vars[12])
		qty_entry.grid(row=8, column=3, sticky="w", padx=10, pady=10)
		
		# button to create and save PO
		submit_btn = tk.Button(t, text="Create PO", command=lambda: self.save_PO(t))
		submit_btn.grid(row=9, column=3, sticky="e", padx=10, pady=10)
	
	# # #
	# Method: add_contract
	# Input: n/a
	# Utility:
	#   Create and open form for adding new contract to backend
	# # #
	def add_contract(self):
		
		# create main contract addition window
		t = tk.Toplevel()
		t.geometry('350x475')
		t.title("Add New Contract")
		
		# instantiate all contract form reference variables
		varList = []
		for i in range(0,10):
			temp = tk.StringVar()
			varList.append(temp)
		
		# label for date awarded
		date_label = tk.Label(t, text="Data Awarded:")
		date_label.grid(row=0, column=0, sticky="e", padx=10, pady=10)
		
		# text box for date awarded
		date_entry = tk.Entry(t, width=12, textvariable=varList[0])
		date_entry.grid(row=0, column=1, sticky="w", padx=10, pady=10)
		
		# label for contract number
		cnumber_label = tk.Label(t, text="Contract #:")
		cnumber_label.grid(row=1, column=0, sticky="e", padx=10, pady=10)
		
		# text box for contract number
		cnumber_entry = tk.Entry(t, width=18, textvariable=varList[1])
		cnumber_entry.grid(row=1, column=1, sticky="w", padx=10, pady=10)
		
		# label for quantity
		qty_label = tk.Label(t, text="Quantity:")
		qty_label.grid(row=2, column=0, sticky="e", padx=10, pady=10)
		
		# text box for quantity
		qty_entry = tk.Entry(t, width=5, textvariable=varList[2])
		qty_entry.grid(row=2, column=1, sticky="w", padx=10, pady=10)
		
		# label for contract total
		ctotal_label = tk.Label(t, text="Contract Total:")
		ctotal_label.grid(row=3, column=0, sticky="e", padx=10, pady=10)
		
		# text box for contract total
		ctotal_entry = tk.Entry(t, width=11, textvariable=varList[3])
		ctotal_entry.grid(row=3, column=1, sticky="w", padx=10, pady=10)
		
		# label for NSN
		nsn_label = tk.Label(t, text="NSN:")
		nsn_label.grid(row=4, column=0, sticky="e", padx=10, pady=10)
		
		# text box for NSN
		nsn_entry = tk.Entry(t, width=15, textvariable=varList[4])
		nsn_entry.grid(row=4, column=1, sticky="w", padx=10, pady=10)
		
		# label for general part description
		partname_label = tk.Label(t, text="Part Name:")
		partname_label.grid(row=5, column=0, sticky="e", padx=10, pady=10)
		
		# text box for general part description
		partname_entry = tk.Entry(t, width=30, textvariable=varList[5])
		partname_entry.grid(row=5, column=1, sticky="w", padx=10, pady=10)
		
		# label for vendor name
		vendor_label = tk.Label(t, text="Vendor Name:")
		vendor_label.grid(row=6, column=0, sticky="e", padx=10, pady=10)
		
		# text box for vendor name
		vendor_entry = tk.Entry(t, width=25, textvariable=varList[6])
		vendor_entry.grid(row=6, column=1, sticky="w", padx=10, pady=10)
		
		# label for part number
		pn_label = tk.Label(t, text="Part #:")
		pn_label.grid(row=7, column=0, sticky="e", padx=10, pady=10)
		
		# text box for part number
		pn_entry = tk.Entry(t, width=20, textvariable=varList[7])
		pn_entry.grid(row=7, column=1, sticky="w", padx=10, pady=10)
		
		# label for preservation method
		preservation_label = tk.Label(t, text="Preservation Method:")
		preservation_label.grid(row=8, column=0, sticky="e", padx=10, pady=10)
		
		# text box for preservation method
		preservation_entry = tk.Entry(t, width=5, textvariable=varList[8])
		preservation_entry.grid(row=8, column=1, sticky="w", padx=10, pady=10)
		
		# label for due date
		date_label = tk.Label(t, text="Due Date:")
		date_label.grid(row=9, column=0, sticky="e", padx=10, pady=10)
		
		# text box for due date
		date_entry = tk.Entry(t, width=12, textvariable=varList[9])
		date_entry.grid(row=9, column=1, sticky="w", padx=10, pady=10)
		
		# button to submit data for contract addition
		submit_button = tk.Button(t, text="Submit Data", command=lambda: self.process_addition(varList,t))
		submit_button.grid(row=10, column=1, sticky="w", padx=10, pady=10)
	
	# # #
	# Method: email_PO
	# Input: n/a
	# Utility:
	#   Create and open form to send PO out to particular vendor
	# # #
	def email_PO(self):
	
		# create main PO email form window
		t = tk.Toplevel()
		t.title("Send PO")
		t.geometry('400x175')
		
		self.current_company.set("") # reset reference variable for selected company
		po_email = tk.StringVar()    # reference variable for selected company's email
		po_display = tk.StringVar()  # cleaned up form of selected PO file
		po_path = tk.StringVar()	 # full path to selected PO file
		
		# update recepient email when company is selected
		def company_function(eventObject):
			company_info = self.PO_dict[self.current_company.get()]
			po_email.set(company_info['email'])
		
		# prompt user to select PO file to be sent, update related variables
		def browse():
			po_path.set(filedialog.askopenfilename())
			po_display.set(po_path.get()[po_path.get().rfind("/")+1:po_path.get().rfind(".")])
		
		# create and send email with supplied information, then close window
		def send_email(window):
			
			msg = MIMEMultipart('related') # create main email object

			msg['From'] = MY_ADDRESS
			msg['To'] = po_email.get()
			# msg['To'] = "k.cook2499@gmail.com"
			msg['Subject'] = "PO"

			msgBody = MIMEMultipart()
			msg.attach(msgBody)

			# populate body of email with template 
			with open('po_email.txt','r') as file:
				msgBody.attach(MIMEText(file.read(),'html'))
			
			# attach selected file to email
			file_name = os.path.basename(po_path.get())
			file = open(po_path.get(),"rb")
			attach = MIMEBase('application', 'octet-stream')
			attach.set_payload((file).read())
			encoders.encode_base64(attach)
			attach.add_header('Content-Disposition', "attachment; filename= %s" % file_name)
			
			# add company logo to email
			try:
				fp = open('logo.png','rb')
				img = MIMEImage(fp.read())
				fp.close()
				img.add_header('Content-ID', '<logo>')
				msg.attach(img)
			except:
				pass
			
			msgBody.attach(attach)
			
			# create connection to email server
			s = smtplib.SMTP(host='smtp-mail.outlook.com', port=587)
			s.starttls()
			
			# login and send email 
			s.login(MY_ADDRESS,MY_PASSWORD)
			s.send_message(msg)
			
			# confirmation message that PO has been sent out
			tk.messagebox.showinfo("PO Sent", "PO has been sent to {}.".format(self.current_company.get()))
			
			window.destroy() # close window
		
		# dropdown menu for vendor names
		company_menu = ttk.Combobox(t, textvariable=self.current_company, values=self.company_list)
		company_menu.bind('<<ComboboxSelected>>', company_function)
		company_menu.grid(row=0, column=0, sticky="we", columnspan=2, padx=10, pady=10)
		
		# label for recipient address
		recipient_label = tk.Label(t, text="Email:")
		recipient_label.grid(row=1, column=0, padx=10, pady=10)
		
		# text box for recipient address
		recipient_entry = tk.Entry(t, width=30, textvariable=po_email)
		recipient_entry.grid(row=1, column=1, padx=10, pady=10)
		
		# label for PO file name
		po_label = tk.Label(t, text="PO File")
		po_label.grid(row=2, column=0, padx=10, pady=10)
		
		# text box for PO file name
		po_entry = tk.Entry(t, width=30, textvariable=po_display)
		po_entry.grid(row=2, column=1, padx=10, pady=10)
		
		# button to select PO file
		po_browse = tk.Button(t, text="Browse", command=browse)
		po_browse.grid(row=2, column=2, padx=10, pady=10)
		
		# button to send out email
		send_email_button = tk.Button(t, text="Send PO", command=lambda: send_email(t))
		send_email_button.grid(row=3, column=1, sticky="e", padx=10, pady=10)
	
	# # #
	# Method: contract_window
	# Input: n/a
	# Utility:
	#   Create main window to select different functions that relate to contracts
	# # #
	def contract_window(self):
		
		# create main contract management window
		main_window = tk.Toplevel()
		main_window.geometry('300x200')
		main_window.title('Contract Management')
		
		def _delete_window():
			try:
				self.save_changes()
				main_window.destroy()
			except:
				pass
				
		main_window.protocol("WM_DELETE_WINDOW", _delete_window)
		
		# open relevant workbooks and create reference dictionaries if config dictionary exists
		t_path = Path('config_dict.json')
		if t_path.is_file():
			self.dict = json.load(open('config_dict.json'))
			self.main_wb = openpyxl.load_workbook(self.dict['main'])
			self.next_row = self.main_wb['DLAORDERS'].max_row+1
			self.next_ref = self.main_wb['DLAORDERS']['A'+str(self.next_row-1)].value+1
			self.wip_wb = openpyxl.load_workbook(self.dict['wip'])
			self.create_dicts()
		
		# button to process new contracts
		add_button = tk.Button(main_window, text="Add New Contract", command= self.add_contract)
		add_button.grid(row=0, column=0, padx=10, pady=10)
		
		# button to create PO
		create_PO_button = tk.Button(main_window, text="Create PO", command= self.create_PO)
		create_PO_button.grid(row=1, column=0, padx=10, pady=10)
		
		# button to send out PO to vendors
		send_PO_button = tk.Button(main_window, text="Send PO", command= self.email_PO)
		send_PO_button.grid(row=2, column=0, padx=10, pady=10)
		
		# button to save changes to main workbook
		save_button = tk.Button(main_window, text="Save Changes", command= self.save_changes)
		save_button.grid(row=3, column=0, padx=10, pady=10)
