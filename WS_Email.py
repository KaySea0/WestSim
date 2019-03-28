import tkinter as tk
import openpyxl
import smtplib
import json
import datetime
from tkinter import filedialog
from tkinter import messagebox
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from string import Template
from pathlib import Path
from settings import *

class WS_Email(object):

	def __init__(self):
	
		self.ali_sheet_name = tk.StringVar() # ALICORP workbook name that is used to create email list
		self.email_list = [] # store list of confirmed emails to be sent out
		self.ws = 0 # monitor width - used to configure where windows load
		self.hs = 0 # monitor height - used to configure where windows load
		
	# # # 
	# Method: read_template
	# Input: 
	#   filename - file that contains template for email w/ html signature
	# Utility:
	#   Open up supplied template file and create object to perform substitutions on
	# Output:
	#   Template object from supplied file
	# # #
	def read_template(self,filename):
	
		with open(filename, 'r',) as template_file:
			template_content = template_file.read()
		return Template(template_content)	
	
	# # #
	# Method: send_emails
	# Input: n/a
	# Utility:
	#   Create email connection object and send all approved quote emails stored in the class' email list
	# # #
	def send_emails(self):
	
		# create email connection
		s = smtplib.SMTP(host='smtp-mail.outlook.com', port=587)
		s.starttls()
		
		s.login(MY_ADDRESS,MY_PASSWORD) # login using defined credentials
		
		# send all emails in approved list
		for msg in self.email_list:
			s.send_message(msg)
	
	# # #
	# Method: process_ali_sheet
	# Input:
	#   prev_window - reference used in order to clear 'browse' window before 'email confirmation' window opens
	# Utility:
	#   (1) Process ALICORP spreadsheet, create potential quote emails to be sent to vendors
	#      (1a) Concurrently create and populate bid sheet for later use when receiving responses / submitting bids
	#   (2) Have user run through all generated emails and confirm validity of each
	# # #
	def process_ali_sheet(self, prev_window):
	
		self.email_list.clear()
	
		# open selected ALICORP sheet to proper sheet
		ali_wb = openpyxl.load_workbook(self.ali_sheet_name.get())
		ali_ws = ali_wb.active
		ali_max_row = ali_ws.max_row
		
		# grab current date in (yyyy-mm-dd) format for bid sheet naming 
		cur_DT = str(datetime.datetime.now())
		cur_date = cur_DT[:cur_DT.find(' ')]
		
		# open config dict and set reference to bid sheet folder
		t_path = Path('config_dict.json')
		bid_folder = None
		if t_path.is_file():
			t_dict = json.load(open('config_dict.json'))
			bid_folder = t_dict['bid']
			
		# open new bid sheet and initialize required vars for editing
		bid_wb = openpyxl.Workbook()
		bid_ws = bid_wb.active
		bid_row = 1
		
		# dictionary contains info relating cage code of vendor to primary email to send quotes to 
		with open('cage_dict.json') as f:
			cage_dict = json.load(f)
			
		email_body = self.read_template('base_email.txt')
		message_list = []   # temp list of all possible emails, require later validation
		part_info = ""      # temp variable to grab info of part(s) for each email
		past_cage_code = "" # keep track of prior cage code to determine grouping / email logic
		
		# # #
		# Notes:
		#   - if cage_dict.get(cage_code,"0") == 0, then the company that is referenced does not currently have a primary email configured
		#     and therefore will not have a message sent (so no processing in this logic block will occur)
		#
		#   - in dictionary entry, 'options' can contain either 'p','g','pg', or nothing
		#     - 'p' = vendor allows for price breaks for a higher quantity
		#     - 'g' = vendor allows for one email to ask for a quote of multiple parts
		# # #
		
		# if ALICORP sheet has only one entry, then run through all steps in one swoop
		if ali_max_row == 2 and cage_dict.get(str(ali_ws['V2'].value),"0") != "0":
			
			# obtain part number (P/N) and quantity (QTY) of desired part
			part_info = "P/N: " + str(ali_ws['X2'].value) + "<br> QTY: " + str(ali_ws['I2'].value) 
			
			# store company name (A), P/N (B), and QTY (C) in next available bid sheet row
			bid_ws['A' + str(bid_row)] = str(ali_ws['W2'].value)
			bid_ws['B' + str(bid_row)] = str(ali_ws['X2'].value)
			bid_ws['C' + str(bid_row)] = str(ali_ws['I2'].value)
			
			# save bid sheet
			bid_wb.save(bid_folder + '/' + cur_date + '_Bid_Sheet.xlsx')
			
			# if vendor allows for price break, add this request to quote
			if 'p' in cage_dict[str(ali_ws['V2'].value)]['options']:
				part_info += " or next price break"
			part_info += "<br><br>"
			
			sub_message = email_body.substitute(PART_INFO = part_info) # using template, generate email body with part_info substitution
			to_address = cage_dict[str(ali_ws['V2'].value)]['email'] # grab vendor contact info from dictionary
			message_list.append((sub_message,to_address)) # add email to list for further validation
			
		else: # for normal execution (ie. more than one row of data)
			for i in range(2,ali_max_row+1):
				
				# temp variables used to make following code cleaner
				cur_vendor_name = str(ali_ws['W' + str(i)].value)
				cur_cage_code = str(ali_ws['V' + str(i)].value)
				cur_PN = str(ali_ws['X' + str(i)].value)
				cur_QTY = str(ali_ws['I' + str(i)].value)
				
				# first line does not have any prior cage codes to compare to for grouping
				if i == 2 and cage_dict.get(cur_cage_code,"0") != "0":
					part_info += "P/N: " + cur_PN + "<br> QTY: " + cur_QTY
					
					bid_ws['A' + str(bid_row)] = cur_vendor_name
					bid_ws['B' + str(bid_row)] = cur_PN
					bid_ws['C' + str(bid_row)] = cur_QTY
					bid_row += 1 # move to next row of bid sheet
					
					if 'p' in cage_dict[cur_cage_code]['options']:
						part_info += " or next price break"
						
					part_info += "<br><br>"
					
					# keep track of prior cage code to determine if prior info chunk needs to be cut off and attached to an email
					past_cage_code = cur_cage_code 
					
				# if not first line and current line has valid contact info...
				elif cage_dict.get(cur_cage_code,"0") != "0":
					
					# if two consecutive lines are from the same company and allow for grouping of quotes...
					if cur_cage_code == past_cage_code and 'g' in cage_dict[cur_cage_code]['options']:
						
						# ...add part info to previous chunk
						part_info += "P/N: " + cur_PN + "<br> QTY: " + cur_QTY
						
						bid_ws['A' + str(bid_row)] = cur_vendor_name
						bid_ws['B' + str(bid_row)] = cur_PN
						bid_ws['C' + str(bid_row)] = cur_QTY
						bid_row += 1
						
						if 'p' in cage_dict[cur_cage_code]['options']:
							part_info += " or next price break"
							
						part_info += "<br><br>"
						
					else: # if current line's cage code does not match prior...
						
						# if prior line is from company with valid contact info...
						if cage_dict.get(past_cage_code,"0") != "0":
						
							# ...cut prior chunk of info and create email
							sub_message = email_body.substitute(PART_INFO = part_info)
							to_address = cage_dict[past_cage_code]['email']
							message_list.append((sub_message,to_address))
						
						past_cage_code = cur_cage_code
						part_info = "P/N: " + cur_PN + "<br> QTY: " + cur_QTY
						
						bid_ws['A' + str(bid_row)] = cur_vendor_name
						bid_ws['B' + str(bid_row)] = cur_PN
						bid_ws['C' + str(bid_row)] = cur_QTY
						bid_row += 1
						
						if 'p' in cage_dict[cur_cage_code]['options']:
							part_info += " or next price break"
						
						part_info += "<br><br>"
					
					# if current chunk of info is from last line, send email with that info
					if i == ali_max_row:
						sub_message = email_body.substitute(PART_INFO = part_info)
						to_address = cage_dict[past_cage_code]['email']
						message_list.append((sub_message,to_address))
				
				# if not first line, not a company with valid contact info, and previous chunk contains data...
				elif part_info != "":
					
					# ...send email with previous chunk
					sub_message = email_body.substitute(PART_INFO = part_info)
					to_address = cage_dict[past_cage_code]['email']
					message_list.append((sub_message,to_address))
					
					part_info = "" # no new info, open to changes when company with valid contact info comes up
					past_cage_code = cur_cage_code
					
		# save bid sheet once done running through ALICORP sheet
		bid_wb.save(bid_folder + '/' + cur_date + '_Bid_Sheet.xlsx')
		
		count = {'value': 0} # keep track of where in message_list user is when validating
	
		# close 'browse' window and open new 'email confirmation' window
		prev_window.destroy()
		t = tk.Toplevel()
		t.title("Email Confirmation")
		
		# set loading location of window to top-left corner
		w = 500
		h = 600
		
		x = (self.ws/4) + 20
		y = (self.hs/4) + 20
		
		t.geometry('%dx%d+%d+%d' % (w,h,x,y))
		
		# configuration options to have text widget display properly
		t.rowconfigure(0,weight=1)
		t.columnconfigure(0,weight=1)
		
		# frame that contains email preview text widget and scrollbar
		email_frame = tk.Frame(t)
		email_frame.grid_rowconfigure(0,weight=1)
		email_frame.grid_columnconfigure(0,weight=1)
		email_frame.grid(row=0,column=0)
		
		# create vertical scrollbar for email preview
		scroll = tk.Scrollbar(email_frame, orient="vertical")
		scroll.grid(row=0,column=1,sticky="ns")
		
		# create text widget for email preview
		email_preview = tk.Text(email_frame,height=500,width=75,yscrollcommand=scroll.set)
		email_preview.insert("1.0","TO: " + message_list[count['value']][1] + "\n" + message_list[count['value']][0].replace("<br><br>","\n").replace("<br>","\n"))
		email_preview.grid(row=0,column=0,sticky="nwe",padx=10,pady=10)
		
		scroll.config(command=email_preview.yview)
		
		# frame for all other UI elements
		user_frame = tk.Frame(t)
		user_frame.grid(row=1,column=0,sticky="sw",padx=5,pady=5)
		
		confirmation_label = tk.Label(user_frame,text="Does this email look correct?")
		confirmation_label.grid(row=0,column=0,sticky="w")
		
		# # #
		# Method: confirm_email
		# Input:
		#   cleared - boolean to determine if current email should be added to validated list for sending
		# Utility:
		#   (1) If email has been validated by user, create message object and store in validated list
		#   (2) Once all emails have been processed, call method that sends all emails
		# # #
		def confirm_email(cleared):
		
			if cleared:
				
				# create message object
				msg = MIMEMultipart('related')

				# assign email properties
				msg['From'] = MY_ADDRESS
				msg['To'] = message_list[count['value']][1]
				# msg['To'] = "k.cook2499@gmail.com"
				msg['Subject'] = "Quote"

				# attach body of email to message object
				msgBody = MIMEMultipart()
				msg.attach(msgBody)

				# enter email text into body
				msgBody.attach(MIMEText(message_list[count['value']][0],'html'))
				
				# add email to validated list
				self.email_list.append(msg)
			
			# move on to next email in need of validation
			count['value'] += 1
			
			# if there are more email to be processed...
			if count['value'] != len(message_list):
				
				# ...load in preview of next email 
				email_preview.delete('1.0',tk.END)
				email_preview.insert("1.0","TO: " + message_list[count['value']][1] + "\n" + message_list[count['value']][0].replace("<br><br>","\n").replace("<br>","\n"))
			else: # if last email has been processed...
			
				# ...send all emails, destroy 'email preview' window, and send confirmation alert
				self.send_emails()
				t.destroy()
				tk.messagebox.showinfo("Email Confirmation", str(len(self.email_list)) + " emails have been sent!")
		
		# set up confirm/reject button 
		confirm_button = tk.Button(user_frame,text="Yes",command=lambda: confirm_email(True))
		confirm_button.grid(row=0,column=1,sticky="w",padx=5)
		
		reject_button = tk.Button(user_frame,text="No",command=lambda: confirm_email(False))
		reject_button.grid(row=0,column=2,sticky="w",padx=5)
	
	# # #
	# Method: email_window
	# Input: n/a
	# Utility:
	#   Open 'browse' window that prompts user to select ALICORP spreadsheet to be processed
	# # #
	def email_window(self):
	
		# open 'browse' window and orient in top left corner of screen
		t = tk.Toplevel()
		t.title("Send Quote Emails")
		
		w = 400
		h = 125
		
		x = (self.ws/4) + 20
		y = (self.hs/4) + 20
		
		t.geometry('%dx%d+%d+%d' % (w,h,x,y))
		
		intro_label = tk.Label(t, text="Select ALICORP spreadsheet you wish to process:")
		intro_label.grid(row=0,column=0,padx=5,pady=10)
		
		# disable changes to entry to display name of selected sheet
		sheet_name_text = tk.Entry(t, state="disabled", textvariable=self.ali_sheet_name, width=50)
		sheet_name_text.grid(row=1,column=0,padx=10,pady=10,sticky="w")
		
		# disable button until user has selected sheet
		process_button = tk.Button(t, text="Process Spreadsheet", state="disabled", command=lambda: self.process_ali_sheet(t))
		process_button.grid(row=2,column=0,padx=5)
		
		# # #
		# Method: browse_function
		# Input: n/a
		# Utility:
		#   Open browse dialog to select ALICORP sheet and, once selected, enable process button
		# # #
		def browse_function():
			self.ali_sheet_name.set(filedialog.askopenfilename())
			process_button.configure(state="normal")
		
		browse_button = tk.Button(t, text="Browse", command = browse_function)
		browse_button.grid(row=1,column=1,padx=5,sticky="w")
		