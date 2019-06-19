import tkinter as tk
from pathlib import Path
from tkinter import messagebox
import json
import openpyxl
import os
import datetime

class WS_Shipping(object):
	
	def __init__(self):
	
		self.idempot = False				  #	boolean signifying that element has been opened before or not
	
		self.contract_list = [] 			  # list of contracts in main_wb - DLAORDERS
		self.wip_list = []					  # list of contracts in wip_wb
		self.ship_inv_list = {}				  # list of contracts in main_wb - ShipInvoice
		
		self.dict = None					  # config dictionary reference
		self.main_wb = None					  # main workbook that contains main contract / shipment information
		self.wip_wb = None					  # secondary workbook that contains detailed information on select contracts
		
		self.contract_var = tk.StringVar()    # reference for contract number for searching
		self.po_var = tk.StringVar()	      # reference for PO number for searching
		self.next_rfid = tk.StringVar()	   	  # reference for next RFID number for WAWF window
		self.shipping_number = tk.StringVar() # reference for shipping number of shipment for WAWF window
		self.invoice_number = tk.StringVar()  # reference for invoice number of shipment for WAWF window
		
		self.queue = []						  # list of contract that user has put into session queue
		
		self.search_frame = None			  # frame that houses search results
		self.canvas = None					  # canvas used to create result pane
		self.scroll = None					  # scrollbar for search results
		
		self.ws_edits = []					  # list that contains shipment information for later saving
		self.next_row = 0 				      # next row number to add data in main_wb
		self.next_ref = 0                     # next reference number for data in main_wb
	
	# # #
	# Method: update_rfid
	# Input: n/a
	# Utility:
	#   Determine last used RFID number in shipping information and increment local reference for future shipments 
	# # #
	def update_rfid(self, update):
	
		# if first time in method, look in workbook for most recent rfid number that was used and set reference
		if self.next_rfid.get() is "":
			
			# open workbook and get max row
			main_ws = self.main_wb['SHipInvice']
			row_search = self.next_row-1
			
			# determine what highest row actually has data (and not just a blank row)
			while(main_ws['K'+str(row_search)].value is None or len(main_ws['K'+str(row_search)].value) < 4): row_search -= 1
			
			# set initial rfid number
			self.next_rfid.set(main_ws['K'+str(row_search)].value)
			
		# if rfid was used in last shipment, update number
		if update:
				
			# increment last 4 digits of rfid number by 1, add leading 0's if necessary
			hex_part = self.next_rfid.get()[-4:]
			hex_int = int(hex_part, 16) + 1
			new_hex = hex(hex_int)[2:].upper()
			if len(new_hex) < 4: new_hex = "0" + str(new_hex)
			
			# set new rfid number
			self.next_rfid.set(self.next_rfid.get()[:-4] + new_hex)
	
	# # #
	# Method: update_view
	# Input: contract number and PO number (pulled from overhead StringVar's)
	# Utility:
	#   Update search result screen with matching contracts based on contract and/or PO number
	# # #
	def update_view(self, queue):
		
		# clear current search results if they exist
		if self.search_frame is not None:
			self.search_frame.destroy()
		
		# local list that contains contracts that will be displayed to user
		display_list = []
		
		# create new frame reference
		self.search_frame = tk.Frame(self.canvas)
		
		# search result header label for contract number
		contract_search_label = tk.Label(self.search_frame, text="Contract #")
		contract_search_label.grid(row=0, column=0, padx=5, pady=5)
		
		# search result header label for PO number
		po_search_label = tk.Label(self.search_frame, text="PO #")
		po_search_label.grid(row=0, column=1, padx=5, pady=5)
		
		# search result header label for vendor name
		vendor_search_label = tk.Label(self.search_frame, text="Vendor")
		vendor_search_label.grid(row=0, column=2, padx=5, pady=5)
		
		# search result header label for quantity
		qty_search_label = tk.Label(self.search_frame, text="Quantity")
		qty_search_label.grid(row=0, column=3, padx=5, pady=5)
		
		# if queue has contracts stored, display them
		if queue and self.queue: display_list = self.queue.copy()
		elif not queue: # if normal search button has been pressed, perform certain search
		
			# if 'search' button is pressed with no inputs, grab all possible contracts
			if not self.contract_var.get() and not self.po_var.get(): display_list = self.contract_list.copy()
			else: 	
				
				# looping through all contracts pulled from main_wb
				for contract in self.contract_list:
				
					# fine tune user contract number search term for proper comparison
					if self.contract_var.get():
						if not self.contract_var.get().isalpha(): contract_comp = self.contract_var.get()
						else: contract_comp = self.contract_var.get().lower()

					# fine tune user po number search term for proper comparison
					if self.po_var.get():
						if not self.po_var.get().isalpha(): po_comp = self.po_var.get()
						else: po_comp = self.po_var.get().lower()
					
					# if user provided input for contract AND PO number...
					if self.contract_var.get():
						if self.po_var.get():
							# check to see if contract matches both values; if so, add to list for later displaying
							if contract_comp in contract[0].lower() and po_comp in contract[1].lower():
								display_list.append(contract)
						else: # if user provided input for only contract number...
							if contract_comp in contract[0].lower():
								display_list.append(contract)
					else: # if user provided input for only PO number...
						if po_comp in contract[1].lower():
							display_list.append(contract)
		
		# loop through all contracts that match search values
		for i in range(1, len(display_list)+1):
			
			contract = display_list[i-1] # easy local reference for displaying
			
			# label for matching contract number
			contract_single_label = tk.Label(self.search_frame, text=contract[0])
			contract_single_label.grid(row=i, column=0, padx=5, pady=5)
			
			# label for matching PO number
			po_single_label = tk.Label(self.search_frame, text=contract[1])
			po_single_label.grid(row=i, column=1, padx=5, pady=5)
			
			# label for matching vendor name
			vendor_single_label = tk.Label(self.search_frame, text=contract[2])
			vendor_single_label.grid(row=i, column=2, padx=5, pady=5)
			
			# label for matching quantity
			qty_single_label = tk.Label(self.search_frame, text=contract[4])
			qty_single_label.grid(row=i, column=3, padx=5, pady=5)
			
			# button that will provide data used in VSM form
			vsm_button = tk.Button(self.search_frame, text="VSM", command=lambda contract=contract: self.vsm_window(contract))
			vsm_button.grid(row=i, column=4, padx=5, pady=5)
			
			# button that will provide data used in WAWF form
			wawf_button = tk.Button(self.search_frame, text="WAWF", command=lambda contract=contract: self.wawf_window(contract))
			wawf_button.grid(row=i, column=5, padx=5, pady=5)
			
			# checkbox that signifies if contract has prior shipment information
			check_qty = (self.ship_inv_list[contract[0]] if contract[0] in self.ship_inv_list else 0)
			ship_check = tk.Checkbutton(self.search_frame, text=str(check_qty)+"/"+str(contract[4]))

			if check_qty != 0: ship_check.select()
			else: ship_check.deselect()
			ship_check.configure(state="disabled")
			
			ship_check.grid(row=i, column=6, padx=5, pady=5, sticky="w")
		
		self.canvas.create_window((4,4),window=self.search_frame,anchor="nw")
		
		self.canvas.update_idletasks()
		self.canvas.configure(scrollregion=self.canvas.bbox('all'), yscrollcommand=self.scroll.set)
		
		self.canvas.pack(side="left",fill="both",expand=True)
		self.scroll.pack(side="right",fill="y")
	
	# # #
	# Method: vsm_window
	# Input:
	#   contract - contract reference that contains pertinent information for form
	# Utility:
	#   Create and display window that provides user necessary information (if available) to complete forms on VSM site
	# # #
	def vsm_window(self, contract):
		
		# create main window
		vsm = tk.Toplevel()
		vsm.title('VSM - ' + contract[0])
		
		# vendor label for current contract 
		company_name = tk.Label(vsm, text=contract[2], font='Helvetica 15 bold')
		company_name.grid(row=0, column=0, sticky="e", padx=5, pady=5)
		
		# get date when contract was saved for reference when saving shipping form from VSM
		if isinstance(contract[3], str):
			date_text = contract[3]
		else:
			date_text = contract[3].strftime('%m/%d/%Y')
		
		# date label for current contract
		contract_date = tk.Label(vsm, text=date_text, font='Helvetica 15 bold')
		contract_date.grid(row=0, column=1, sticky="w", padx=5, pady=5)
		
		# label for textbox that gives first 6 digits of contract number to use in search
		po_label = tk.Label(vsm, text="PO Search:")
		po_label.grid(row=1, column=0, sticky="e", padx=5, pady=5)
		
		# textbox that contains first 6 digits of current contract number for searching
		po_search = tk.Text(vsm, height=1, width=8, borderwidth=0)
		po_search.insert(1.0, contract[0][:6])
		po_search.configure(state="disabled", inactiveselectbackground=po_search.cget("selectbackground"))
		po_search.grid(row=1, column=1, sticky="w", padx=5, pady=5)
		
		# label that shows last 4 digits of contract number for quick reference when searching
		po_ref = tk.Label(vsm, text="-" + contract[0][contract[0].rfind('-')+1:])
		po_ref.grid(row=1, column=2, padx=5, pady=5)
		
		# variables that will store part number and preservation method of contract if found in wip_wb
		pn_text = ""
		preservation_text = ""
		
		# look through all contracts in wip_wb
		for ref in self.wip_list:
			# if reference is found, pull part number and preservation method
			if ref[0] == contract[0]:
				pn_text = ref[2]
				preservation_text = ref[3]
		
		# if information is not found in wip_wb, put default message
		if not pn_text: 
			pn_text = "Look up in contract PDF"
			preservation_text = "Look up in contract PDF"
		
		# label for textbox that gives part number of contract
		pn_label = tk.Label(vsm, text="P/N:")
		pn_label.grid(row=2, column=0, sticky="e", padx=5, pady=5)
		
		# textbox that contains part number of contract (if available)
		pn_info = tk.Text(vsm, height=1, width=20, borderwidth=0)
		pn_info.insert(1.0, pn_text)
		pn_info.configure(state="disabled", inactiveselectbackground=po_search.cget("selectbackground"))
		pn_info.grid(row=2, column=1, sticky="w", padx=5, pady=5)
		
		# label for textbox that gives preservation method of contract
		preservation_label = tk.Label(vsm, text="Preservation Method:")
		preservation_label.grid(row=3, column=0, sticky="e", padx=5, pady=5)
		
		# textbox that contains preservation method of contract (if available)
		preservation_info = tk.Text(vsm, height=1, width=20, borderwidth=0)
		preservation_info.insert(1.0, preservation_text)
		preservation_info.configure(state="disabled", inactiveselectbackground=po_search.cget("selectbackground"))
		preservation_info.grid(row=3, column=1, sticky="w", padx=5, pady=5)
		
		# button that will add contract to session queue
		vsm_queue = tk.Button(vsm, text="Add to Queue", command=lambda: self.queue.append(contract))
		vsm_queue.grid(row=4, column=2, padx=5, pady=5)
	
	# # #
	# Method: wawf_window
	# Input:
	#   contract - contract reference that contains pertinent information for form
	# Utility:
	#   Create and display form that provides necessary information to complete form on WAWF site
	# # #
	def wawf_window(self, contract):
		
		# create main window
		wawf = tk.Toplevel()
		wawf.title('WAWF - ' + contract[0])
		
		# open appropriate sheet and get max row
		main_ws = self.main_wb['SHipInvice']
		max_row = main_ws.max_row
		
		rfid_display = tk.StringVar()	 # reference for rfid number that will be displayed in window
		total_price = tk.StringVar()	 # reference for calculated total of contract based on quantity
		
		qty_var = tk.IntVar()			 # reference for quantity of item in shipment
		total_var = tk.IntVar()			 # reference for total radiobutton that signifies that total has been verified
		rfid_var = tk.IntVar()			 # reference for rfid checkbox that signifies that rfid will be used
		final_var = tk.IntVar()			 # reference for final shipment checkbox that signifies that shipment will be final one of contract
		
		# reset final shipment checkbox
		final_var.set(0)
		
		# set initial value of rfid display and reset related checkbox
		rfid_display.set("No")
		rfid_var.set(0)
		
		# set initial value of contract total display and reset related radiobutton
		total_var.set(-1)
		total_price.set("$" + str(contract[5]))
		
		# reset shipping number to base value if user closed window before saving
		if 'Z' in self.shipping_number.get():
			self.shipping_number.set(self.shipping_number.get()[:-1])
		
		# # #
		# Method: check_select
		# Utility:
		#   Change shipping number based on "final" designation
		# # #
		def check_select():
			
			# add / remove 'Z' based on final shipment designation
			if 'Z' in self.shipping_number.get():
				self.shipping_number.set(self.shipping_number.get()[:-1])
			else:
				self.shipping_number.set(self.shipping_number.get()+'Z')
				
			update_text(shipping_text, self.shipping_number)
		
		# # #
		# Method: rfid_select
		# Utility:
		#   Change rfid display based on necessity 
		# # #
		def rfid_select():
			
			# swap between "No" and next rfid number 
			if rfid_display.get() == "No":
				rfid_display.set(self.next_rfid.get())
			else:
				rfid_display.set("No")
			
			# update rfid display textbox
			update_text(rfid_text, rfid_display)
		
		# # #
		# Method: qty_change
		# Utility:
		#   Track when quantity is changed, alter total display when appropriate
		# # #
		def qty_change(*args):
			
			# calculate unit price based on displayed total and max quantity
			unit_price = round(contract[5] / contract[4], 2)
			
			# attempt to cast quantity as int; if not possible, default to 0
			try:
				qty = int(qty_var.get())
			except:
				qty = 0
				
			# if quantity is within appropriate range, determine total and display; otherwise, show error
			if qty <= contract[4] and qty > 0: 
				total_price.set("$" + str(format(unit_price * qty, '.2f')))
			else:
				total_price.set("Error")
			
			# update contract total display textbox
			update_text(total_text, total_price)
			
		# # #
		# Method: update_text
		# Utility:
		#   Update provided textbox ('text') with provided data ('var')
		# # #
		def update_text(text, var):
			
			text.configure(state="normal")
			text.delete(1.0, tk.END)
			text.insert(1.0, var.get())
			text.configure(state="disabled")
		
		# # #
		# Method: update_numbers
		# Utility:
		#   Increment shipping and invoice numbers for multiple shipments in one program iteration
		def update_numbers():
			
			# cut down shipping number for uniformity
			if 'Z' in self.shipping_number.get():
				self.shipping_number.set(self.shipping_number.get()[:-1])
			
			# grab part of numbers that will change between shipments
			shipping_partial = self.shipping_number.get()[-3:]
			invoice_partial = self.invoice_number.get()[-4:]
			
			# increment integer parts of partial shipping / invoice numbers
			next_ship_int = int(shipping_partial)+1
			next_inv_int = int(invoice_partial)+1
			
			# stitch together base part and updated part of shipping number
			next_ship = str(next_ship_int)
			next_ship = (3-len(next_ship))*"0" + next_ship
			
			# stitch together base part and updated part of invoice number
			next_inv = str(next_inv_int)
			next_inv = (4-len(next_inv))*"0" + next_inv
			
			# update reference variable of shipping/invoice numbers
			self.shipping_number.set(self.shipping_number.get()[:-3] + next_ship)
			self.invoice_number.set(self.invoice_number.get()[:-4] + next_inv)
			
		# # #
		# Method: add_edit
		# Utility:
		#   Check to see if all data is valid before saving data on shipment for later addition into workbook
		# # #
		def add_edit(window, row):
			
			# attempt to cast quantity as int; if not possible, default to 0
			try:
				qty = int(qty_var.get())
			except:
				qty = 0
				
			# if quantity is not in proper range, set to default value to trigger error
			if qty < 1 or qty > contract[4]: qty = 0
				
			# if quantity is correct and total verification button has been pressed, save data for later addition into workbook
			if qty != 0 and total_var.get() != -1:
				
				# get current date
				current_date = datetime.datetime.now().strftime("%m/%d/%Y")
				
				# save data on shipment to be put into workbook
				# row to add data to - reference number - contract number - vendor name - quantity - total price - shipping number - invoice number - date of shipment - rfid number (when applicable)
				self.ws_edits.append([self.next_row, self.next_ref, contract[0], contract[2], qty, total_price.get(), self.shipping_number.get(), self.invoice_number.get(), current_date, rfid_display.get(), row])
				
				# update all references for future additions
				self.next_row += 1
				self.next_ref += 1
				self.update_rfid(rfid_display.get() != "No")
				update_numbers()
				
				# send confirmation alert before closing window
				tk.messagebox.showinfo("Information Saved", "The information for this contract has been saved!")
				window.destroy()
			
			# if quantity is wrong or total verification has not been confirmed, alert user as such and perform no saving operations
			else:
				
				tk.messagebox.showinfo("An error has been detected", "Information could not be saved due to an uresolved issue with the 'Quantity' section and/or a lack of verification in the 'Total' section. Please look over the section(s) and attempt submitting again.")
		
		qty_var.set(contract[4]) # set default value of quantity to what is listed in workbook
		qty_var.trace("w", qty_change) # set trace method as to update total automatically
		
		# main title of section to fill room
		section_title = tk.Label(wawf, text="WAWF", font='Helvetica 15 bold')
		section_title.grid(row=0, column=0, padx=5, pady=5)
		
		# frame to bundle quantity elements / final shipment checkbox together cleanly
		qty_frame = tk.Frame(wawf)
		
		# checkbox to determine if shipment will be last one of contract
		tk.Checkbutton(qty_frame, text="Final Shipment?", variable = final_var, command=check_select).grid(row=0, column=0, padx=5, pady=5, sticky="e")
		
		# textbox that contains quantity for contract - can be edited by user
		qty_text = tk.Entry(qty_frame, width=5, textvariable=qty_var)
		qty_text.grid(row=0, column=1, padx=5, pady=5, sticky="e")
		
		# label for quantity textbox to show maximum value
		qty_total = tk.Label(qty_frame, text="of " + str(contract[4]))
		qty_total.grid(row=0, column=2, padx=5, pady=5)
		
		# place quantity / final shipment container in proper place
		qty_frame.grid(row=0, column=2, columnspan=3, padx=5, pady=5, sticky='e')
		
		# label for contract number textbox
		contract_label = tk.Label(wawf, text="Contract #:")
		contract_label.grid(row=1, column=0, padx=5, pady=5, sticky="e")
		
		# textbox that contains contract number - cannot be edited by user (but can be selected)
		contract_search = tk.Text(wawf, height=1, width=18, borderwidth=0)
		contract_search.insert(1.0, contract[0])
		contract_search.configure(state="disabled", inactiveselectbackground=contract_search.cget("selectbackground"))
		contract_search.grid(row=1, column=1, padx=5, pady=5, sticky="w")
		
		# label for shipping number textbox
		shipping_label = tk.Label(wawf, text="Shipping #:")
		shipping_label.grid(row=2, column=0, padx=5, pady=5, sticky="e")
		
		# textbox that contains shipping number - cannot be edited by user directly (but can be selected)
		shipping_text = tk.Text(wawf, height=1, width=10, borderwidth=0)
		shipping_text.insert(1.0, self.shipping_number.get())
		shipping_text.configure(state="disabled", inactiveselectbackground=contract_search.cget("selectbackground"))
		shipping_text.grid(row=2, column=1, padx=5, pady=5, sticky="w")
		
		# label for invoice number textbox
		invoice_label = tk.Label(wawf, text="Invoice #")
		invoice_label.grid(row=3, column=0, padx=5, pady=5, sticky="e")
		
		# textbox that contains invoice number - cannot be edited by user (but can be selected)
		invoice_text = tk.Text(wawf, height=1, width=10, borderwidth=0)
		invoice_text.insert(1.0, self.invoice_number.get())
		invoice_text.configure(state="disabled", inactiveselectbackground=contract_search.cget("selectbackground"))
		invoice_text.grid(row=3, column=1, padx=5, pady=5, sticky="w")
		
		# checkbox to determine if RFID number must be used for shipment
		tk.Checkbutton(wawf, text="RFID #:", variable=rfid_var, command=rfid_select).grid(row=2, column=2, padx=5, pady=5, sticky="e")
		
		# textbox that contains RFID number if it is required - cannot be edited by user directly (but can be selected)
		rfid_text = tk.Text(wawf, height=1, width=25, borderwidth=0)
		rfid_text.insert(1.0, rfid_display.get())
		rfid_text.configure(state="disabled", inactiveselectbackground=contract_search.cget("selectbackground"))
		rfid_text.grid(row=2, column=3, padx=5, pady=5)
		
		# radiobutton that must be selected to signify total has been verified between WAWF and contract total
		total_button = tk.Radiobutton(wawf, text="Total:", variable=total_var, value=1)
		total_button.grid(row=1, column=2, padx=5, pady=5, sticky="e")
		
		# textbox that determines calculated total of contract - cannot be edited by user directly
		total_text = tk.Text(wawf, height=1, width=10, borderwidth=0)
		total_text.insert(1.0, "$" + str(contract[5]))
		total_text.configure(state="disabled", inactiveselectbackground=contract_search.cget("selectbackground"))
		total_text.grid(row=1, column=3, padx=5, pady=5, sticky="w")
		
		# frame to bundle submit/cancel buttons together
		button_frame = tk.Frame(wawf)
		
		# button used to call checks and attempt to save shipment information
		submit_button = tk.Button(button_frame, text="Submit", command=lambda: add_edit(wawf, contract[6]))
		submit_button.grid(row=0, column=0, padx=5, pady=5)
		
		# button used to cancel operation if contract was open on accident
		cancel_button = tk.Button(button_frame, text="Cancel", command=lambda: wawf.destroy())
		cancel_button.grid(row=0, column=1, padx=5, pady=5)
		
		# place submit/cancel container in proper place
		button_frame.grid(row=3, column=3, padx=5, pady=5)
	
	# # #
	# Method: save_changes
	# Input: n/a
	# Utility:
	#   Save information about processed shipments to main_wb 
	# # #
	def save_changes(self):
		
		# if any changes have been made, process them
		if self.ws_edits:
			
			# open up sheet that contains shipment information
			main_ws = self.main_wb['SHipInvice']
			
			# running through all contracts to be saved
			for edit in self.ws_edits:
				
				row = str(edit[0]) # row number that new info will be saved on 
				if edit[8][0] == '0': edit[8] = edit[8][1:] # if month is meant to be one digit, shorten date display
				
				main_ws['A'+row] = edit[1] 	# reference number of shipment
				main_ws['B'+row] = edit[2]	# contract number
				main_ws['C'+row] = edit[3]	# vendor name
				main_ws['E'+row] = edit[4]	# quantity
				main_ws['G'+row] = edit[5]	# shipment total
				main_ws['H'+row] = edit[6]	# shipping number
				main_ws['I'+row] = edit[7]	# invoice number
				main_ws['J'+row] = edit[8]	# date of shipment
				main_ws['K'+row] = edit[9]  # RFID number
				
				# make font of quantity column red to match preset
				main_ws['E'+row].font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
				
				# unhighlight contract in DLAORDERS once final shipment has been made
				if edit[6][-1:] == "Z":
					contract_ws = self.main_wb['DLAORDERS']
					normal = openpyxl.styles.Side(border_style="thin", color="C0C0C0")
					
					contract_ws['B'+str(edit[10])].fill = openpyxl.styles.PatternFill("solid", fgColor="FFFFFF")
					contract_ws['B'+str(edit[10])].border = openpyxl.styles.Border(top=normal, left=normal, right=normal, bottom=normal)
			
			# attempt to save changes to workbook
			while True:
				try: # if workbook is closed, changes will be saved
					self.main_wb.save(self.dict['main']) # save main_wb
					break
				except:
					pass
				
				# if workbook is not closed, inform user of this fact and attempt to save again after 'ok' is pressed
				tk.messagebox.showinfo("An error has occurred", "Workbook {} is still open, please close it and press 'Ok' to continue operation.".format(self.dict['main'][self.dict['main'].rfind('/')+1:]))
			
			# clear edit list and show save confirmation message
			self.ws_edits.clear()
			tk.messagebox.showinfo("Changes confirmation", "All changes have been saved to Workbook {}.".format(self.dict['main'][self.dict['main'].rfind('/')+1:]))
		
	# # #
	# Method: create_lists
	# Input: n/a
	# Utility:
	#   Create reference lists for all contracts (from main_wb) and recently added contracts that contain extra information (wip_wb)
	# # #
	def create_lists(self):
	
		self.contract_list.clear()
		self.wip_list.clear()
		self.ship_inv_list.clear()
	
		# if lists have not been created yet, do so
		if not self.contract_list:
			
			ship_inv_ws = self.main_wb['SHipInvice'] # open main_wb - ShipInvoice
			list_end = ship_inv_ws.max_row+1
			
			# # create list of all contract numbers and total associated quantity that exists as shipments in records
			for i in range(1, list_end):
				if not ship_inv_ws['B'+str(i)].value is None and not ship_inv_ws['E'+str(i)].value is None:
					ship_num = ship_inv_ws['B'+str(i)].value
					
					if ship_num in self.ship_inv_list:
						self.ship_inv_list[ship_num] += ship_inv_ws['E'+str(i)].value
					else:
						self.ship_inv_list[ship_num] = ship_inv_ws['E'+str(i)].value
			
			main_ws = self.main_wb['DLAORDERS'] # open main_wb - DLAORDERS
			list_start = 278 # row number that search should start at for main_ws (dependent on what orders have come in)
			list_end = main_ws.max_row+1
			
			# # create list of all contracts from main_wb; data outlined below
			# Contract Number - PO Number - Vendor Name - Date Awarded - Quantity - Contract Total - main_ws row number
			for i in range(list_start, list_end):
				if not main_ws['I'+str(i)].value is None:
					
					data = [main_ws['B'+str(i)].value, main_ws['I'+str(i)].value, main_ws['F'+str(i)].value, main_ws['G'+str(i)].value, main_ws['E'+str(i)].value, main_ws['K'+str(i)].value, i]
					self.contract_list.append(data)
			
			# open wip_wb and get reference for final row
			wip_ws = self.wip_wb.active
			list_end = wip_ws.max_row+1
			
			# # create list of all contracts from wip_wb; data outlined below
			# Contract Number - Vendor Name - P/N - Preservation Method
			for i in range(1, list_end):
				self.wip_list.append([wip_ws['B'+str(i)].value, wip_ws['G'+str(i)].value, wip_ws['H'+str(i)].value, wip_ws['I'+str(i)].value])
	
	# # #
	# Method: shipping_window
	# Input: n/a
	# Utility:
	#   Create main search window for user to get information on contracts for shipping purposes
	# # #
	def shipping_window(self):
		
		# create main shipping information window
		t = tk.Toplevel()
		t.geometry('600x250')
		t.title('Shipping Management')
		
		def _delete_window():
			try:
				self.save_changes()
				t.destroy()
			except:
				pass
				
		t.protocol("WM_DELETE_WINDOW", _delete_window)
		
		# open relevant workbooks and create reference lists/variables if config dictionary exists
		t_path = Path('config_dict.json')
		if t_path.is_file():
			self.dict = json.load(open('config_dict.json'))
			self.main_wb = openpyxl.load_workbook(self.dict['main'])
			self.wip_wb = openpyxl.load_workbook(self.dict['wip'])
			
			# open main_wb and determine last non-empty row in worksheet
			check_row = self.main_wb['SHipInvice'].max_row
			while(self.main_wb['SHipInvice']['A'+str(check_row)].value is None): check_row -= 1
			self.next_row = check_row+1
			
			self.next_ref = self.main_wb['SHipInvice']['A'+str(self.next_row-1)].value+1 # get next reference number for shipment info
			
			# set initial values of invoice and shipment number
			inv_num = int(self.main_wb['SHipInvice']['I'+str(self.next_row-1)].value[-4:])+1
			self.invoice_number.set(self.main_wb['SHipInvice']['I'+str(self.next_row-1)].value[:-4] + (4-len(str(inv_num)))*"0" + str(inv_num))
		
			self.shipping_number.set(self.invoice_number.get()[:-4] + self.invoice_number.get()[-3:])
			
			# create contract lists from main/wip_wb and grab next rfid number
			self.create_lists()
			self.update_rfid(not self.idempot)
			
			if not self.idempot: self.idempot = True
			self.queue.clear()
			
		self.canvas = tk.Canvas(t, borderwidth=0)
		self.scroll = tk.Scrollbar(t, orient="vertical", command=self.canvas.yview)
		
		# frame that contains all user input elements (found above search results area)
		input_frame = tk.Frame(t)
		
		# label for contract number search
		contract_label = tk.Label(input_frame, text="Contract Number:")
		contract_label.grid(row=0, column=0, padx=5, pady=5)
		
		# entry box for contract number search
		contract_entry = tk.Entry(input_frame, width=20, textvariable=self.contract_var)
		contract_entry.grid(row=0, column=1, padx=5, pady=5)
		
		# label for PO number search
		po_label = tk.Label(input_frame, text="PO Number:")
		po_label.grid(row=0, column=2, padx=5, pady=5)
		
		# entry box for PO number search
		po_entry = tk.Entry(input_frame, width=10, textvariable=self.po_var)
		po_entry.grid(row=0, column=3, padx=5, pady=5)
		
		# search button for finding matching contracts
		search_button = tk.Button(input_frame, text="Search", command=lambda: self.update_view(False))
		search_button.grid(row=0, column=4, padx=5, pady=5)
		
		# queue button for pulling up contracts that are being worked on in current session
		queue_button = tk.Button(input_frame, text="Queue", command=lambda: self.update_view(True))
		queue_button.grid(row=0, column=5, padx=5, pady=5)
		
		# pack input elements into window
		input_frame.pack(anchor="n", fill=tk.X)
		